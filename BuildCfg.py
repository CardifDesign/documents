import xlrd
import openpyxl
import Enviroments
import shutil
from datetime import date
import os

#defina variable enve dev,uat,prd
enves = 'prd'
APIMashery = 'benefits'
workpad = "C:\\Users\\d87797\\Documents\\Mashery\\"
plantillaf = "plantilla.xlsx"


CFGFILENAME = "Mashery"+" "+enves.upper()+"_CO_"+APIMashery.upper()+".xlsx"
Sorce = 'source_'+APIMashery.lower()+'.xlsx'
Sheetcfg = enves.lower()


plantilla = workpad+plantillaf
SorceFile = workpad+Sorce
CFGPATH="C:\\Users\\d87797\\Documents\\Mashery\\"+str(date.today())
FileCFG = workpad+CFGFILENAME
FileCFGOK = CFGPATH+"\\"+CFGFILENAME


Mashery_dir = Enviroments.vars.env(enves)[1]
BW_dir = Enviroments.vars.env(enves)[0]
shutil.copy(plantilla, FileCFG)
documento = xlrd.open_workbook(SorceFile)
apis = documento.sheet_by_name(Sheetcfg)


workbook2 = openpyxl.load_workbook(FileCFG)

for i in range(1, apis.nrows):

    fila = apis.row(i)
    api_name_mashery = apis.cell_value(i, 0).strip()
    print(fila)
    #worksheet = workbook.add_worksheet(name=api_name_mashery)
    workbook2.active.title = 'plantilla'
    origen = workbook2['plantilla']
    destino = workbook2.copy_worksheet(origen)
    destino.title = api_name_mashery
    workbook2.save(FileCFG)

workbook2.close()

workbook2 = openpyxl.load_workbook(FileCFG)

for i in range(1, apis.nrows):

    api_name_mashery = apis.cell_value(i, 0).strip()
    urlmashery = apis.cell_value(i, 1).strip()
    portBW = apis.cell_value(i, 2)
    portBW = str(portBW).replace(".0", "").strip()
    bw_metods = apis.cell_value(i, 3).strip()
    suports = apis.cell_value(i, 4).strip()
    end_point_type = apis.cell_value(i, 5).strip()

    full_url_bw = BW_dir+":"+portBW+bw_metods
    full_url_mashery = Mashery_dir+urlmashery
    workbook2.active = i+1
    workbook2.active['E5'] = end_point_type
    workbook2.active['E8'] = api_name_mashery
    workbook2.active['E9'] = full_url_mashery
    workbook2.active['E19'] = suports
    if end_point_type.lower() == 'token':
        workbook2.active['E43'] = 'Enabled'
        workbook2.active['E44'] = 'Enabled'
        workbook2.active['E48'] = ''
        workbook2.active['E10'] = ''
    else:
        workbook2.active['E10'] = full_url_bw
        workbook2.active['E43'] = 'Enabled'


    print(workbook2.active)
    workbook2.save(FileCFG)

workbook2.close()

workbook2 = openpyxl.load_workbook(FileCFG)
del workbook2["plantilla"]
workbook2.save(FileCFG)
workbook2.close()

if not os.path.exists(CFGPATH):
    os.makedirs(CFGPATH)

shutil.copy(FileCFG, FileCFGOK)





